from typing import Optional
import sys
from PySide6 import QtCore
from PySide6.QtCore import QCoreApplication
from PySide6.QtGui import QIcon,QColor,QStandardItemModel, QStandardItem
from PySide6.QtWidgets import (QApplication, QMainWindow,QMessageBox, QWidget)
from ui_main import Ui_MainWindow
from ui_login import Ui_Login
from datetime import datetime,timedelta
import openpyxl
import smtplib
from email.mime.text import MIMEText

class Login(QWidget,Ui_Login):
    def __init__(self) -> None:
        super(Login,self).__init__()
        self.setupUi(self)
        self.setWindowTitle("Login")
        
        self.load_workbook("controle_de_barcos.xlsx")
        
        appIcon = QIcon("./Imagens/barco.png")
        self.setWindowIcon(appIcon)
        
        self.btn_login.clicked.connect(self.verificacao_entrada)
        if self.ultimo_dia_do_mes():
            self.btn_login.clicked.connect(self.verifica_combustivel)
            data_atual = datetime.now()
            data = data_atual.strftime("%d-%m-%Y")
            
            self.workbook.save(f'./BKP/controle_de_barcos-{data}.xlsx')
            self.load_workbook('./BKP/Planilha_Base/controle_de_barcos.xlsx')
            self.workbook.save('controle_de_barcos.xlsx')
            
            
        self.btn_login.clicked.connect(self.verifica_horas)
                      
    def load_workbook(self, filename):
        self.workbook = openpyxl.load_workbook(filename)
        self.barcos_sheet = self.workbook["Barcos"]

    def verificacao_entrada(self):
        if self.senha_lineEdit.text() == '552admin!' and self.usuario_lineEdit.text() == 'admin':
            self.w = MainWindow()
            self.close()
            self.w.show()
            
        else:
            cor_personalizada = QColor(0, 0, 0)
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText(f"Credenciais Incorretas!")
            msg.setStyleSheet(f"QLabel {{ color: {cor_personalizada.name()}; }}")
            msg.setWindowTitle("Código da Viagem")
            msg.exec_()
            self.senha_lineEdit.clear()
            self.usuario_lineEdit.clear()
      
    def enviar_email(self, barco):
        # Configuração do servidor SMTP
        smtp_server = 'smtp.gmail.com'
        smtp_port = 587  
        smtp_username = 'controlebarcos@gmail.com'
        smtp_password = 'sewgsbnxwetbzqps'
        
        # Destinatário
        to_email = 'controlebarcos@gmail.com'

        # Criar mensagem de email
        subject = 'Aviso de Horas do Barco'
        message = f'O barco {barco} atingiu 100 horas de uso. Recomendação trocar o óleo!'

        msg = MIMEText(message)
        msg['Subject'] = subject
        msg['From'] = smtp_username
        msg['To'] = to_email

        # Conectar ao servidor SMTP e enviar email
        try:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.sendmail(smtp_username, to_email, msg.as_string())
            server.quit()
            #print(f'Email enviado para {to_email}')
        except Exception as e:
            #print(f'Erro ao enviar email: {str(e)}')
            pass

    def enviar_email_2(self, barco):
        # Configuração do servidor SMTP
        smtp_server = 'smtp.gmail.com'
        smtp_port = 587  # Porta para TLS
        smtp_username = 'controlebarcos@gmail.com'
        smtp_password = 'sewgsbnxwetbzqps'
        
        # Destinatário
        to_email = 'controlebarcos@gmail.com'

        # Criar mensagem de email
        subject = 'Aviso de Horas do Barco'
        message = f'O barco {barco} atingiu 300 horas de uso. Recomendação é de trocas as velas!'

        msg = MIMEText(message)
        msg['Subject'] = subject
        msg['From'] = smtp_username
        msg['To'] = to_email

        # Conectar ao servidor SMTP e enviar email
        try:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.sendmail(smtp_username, to_email, msg.as_string())
            server.quit()
            #print(f'Email enviado para {to_email}')
        except Exception as e:
            #print(f'Erro ao enviar email: {str(e)}')
            pass
        
    def verifica_horas(self):
        for index, barcos in enumerate(self.barcos_sheet.iter_rows(values_only=True)):
            if index == 0:
                continue

            self.aba_barcos = self.workbook['Barcos']

            if barcos[3] is not None:
                if barcos[3] >= 6000:
                    if barcos[3] > 6000:
                        novo_valor = barcos[3] - 6000
                        self.aba_barcos.cell(row=barcos[0], column=4, value=novo_valor)

                        contador = barcos[4] + 1
                        self.aba_barcos.cell(row=barcos[0], column=5, value=contador)

                        self.enviar_email(barco=barcos[1])
                        self.workbook.save('controle_de_barcos.xlsx')
                        
                        self.load_workbook('./BKP/Planilha_Base/controle_de_barcos.xlsx')
                        self.aba_barcos = self.workbook['Barcos']
                        self.aba_barcos.cell(row=barcos[0], column=4, value=novo_valor)
                        self.aba_barcos.cell(row=barcos[0], column=5, value=contador)
                        self.workbook.save('./BKP/Planilha_Base/controle_de_barcos.xlsx')
                        
                        self.load_workbook('controle_de_barcos.xlsx')
                        self.aba_barcos = self.workbook['Barcos']

                    else:
                        zera_minutos = 0
                        self.aba_barcos.cell(row=barcos[0], column=4, value=zera_minutos)

                        contador = barcos[4] + 1
                        self.aba_barcos.cell(row=barcos[0], column=5, value=contador)

                        self.enviar_email(barco=barcos[1])
                        self.workbook.save('controle_de_barcos.xlsx')
                        
                        self.load_workbook('./BKP/Planilha_Base/controle_de_barcos.xlsx')
                        self.aba_barcos = self.workbook['Barcos']
                        self.aba_barcos.cell(row=barcos[0], column=4, value=zera_minutos)
                        self.aba_barcos.cell(row=barcos[0], column=5, value=contador)
                        self.workbook.save('./BKP/Planilha_Base/controle_de_barcos.xlsx')
                        
                        self.load_workbook('controle_de_barcos.xlsx')
                        self.aba_barcos = self.workbook['Barcos']
                        

                    self.workbook.save('controle_de_barcos.xlsx')
                    print(barcos)

                if barcos[4] >= 3:
                    self.aba_barcos.cell(row=barcos[0], column=5, value=0)
                    self.workbook.save('controle_de_barcos.xlsx')
                    self.enviar_email_2(barco=barcos[1])
                    
                    self.load_workbook('./BKP/Planilha_Base/controle_de_barcos.xlsx')
                    self.aba_barcos = self.workbook['Barcos']
                    self.aba_barcos.cell(row=barcos[0], column=5, value=0)
                    self.workbook.save('./BKP/Planilha_Base/controle_de_barcos.xlsx')
                    
                    self.load_workbook('controle_de_barcos.xlsx')
                    self.aba_barcos = self.workbook['Barcos']

    def ultimo_dia_do_mes(self):
        today = datetime.now()
        next_day = today + timedelta(days=1)

        return today.month != next_day.month

    def verifica_combustivel(self):
        for index, barcos in enumerate(self.barcos_sheet.iter_rows(values_only=True)):
            if index == 0:
                continue
            
            if barcos[1] is not None and barcos[5] is not None:
                self.enviar_email_combustivel(barco=barcos[1], combustivel=barcos[5])

            #self.aba_barcos = self.workbook['Barcos']

    def enviar_email_combustivel(self, barco,combustivel):
        # Configuração do servidor SMTP
        smtp_server = 'smtp.gmail.com'
        smtp_port = 587  # Porta para TLS
        smtp_username = 'controlebarcos@gmail.com'
        smtp_password = 'sewgsbnxwetbzqps'
        
        # Destinatário
        to_email = 'controlebarcos@gmail.com'

        # Criar mensagem de email
        subject = f'{barco} Combustivel'
        
        message = f'O barco {barco} usou {combustivel} litros de combustivel'

        msg = MIMEText(message)
        msg['Subject'] = subject
        msg['From'] = smtp_username
        msg['To'] = to_email

        # Conectar ao servidor SMTP e enviar email
        try:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.sendmail(smtp_username, to_email, msg.as_string())
            server.quit()
            #print(f'Email enviado para {to_email}')
        except Exception as e:
            #print(f'Erro ao enviar email: {str(e)}')
            pass
    
class MainWindow(QMainWindow,Ui_MainWindow):
    
    def __init__(self) :
        super(MainWindow,self).__init__()
        self.setupUi(self)
        self.setWindowTitle("Sistema de Barcos")
        
        self.barcos_sheet = None
        self.pilotos_sheet = None
        self.viagens_sheet = None
        
        self.pilotos_model = QStandardItemModel()
        self.barcos_model = QStandardItemModel() 
        self.pilotos_comboBox.setModel(self.pilotos_model)
        self.barcos_comboBox.setModel(self.barcos_model)
            
        self.load_workbook("controle_de_barcos.xlsx")
            

        #Definindo Icone
        appIcon = QIcon("./Imagens/barco.png")
        self.setWindowIcon(appIcon)
        
        #Botão Menu
        self.btn_toggle.clicked.connect(self.leftMenu)
        
        ######################################################################################################################
        #Navegação
        self.btn_home.clicked.connect(lambda: self.Pages.setCurrentWidget(self.pg_home))
        self.btn_cadastrar.clicked.connect(lambda: self.Pages.setCurrentWidget(self.pg_cadastrar))
        self.btn_consulta.clicked.connect(lambda: self.Pages.setCurrentWidget(self.pg_consulta))
        self.btn_viagens.clicked.connect(lambda: self.Pages.setCurrentWidget(self.pg_viagens))
        
        ######################################################################################################################
        #Funções da Pagina Viagem
        self.carregar_pilotos_combobox() # Carrega Pilotos
        self.carregar_barcos_combobox() #Carrega Barcos
        
        self.btn_enviar.clicked.connect(self.adicionar_viagem) #Envia viagem para a planilha
        
        self.btn_enviar_volta.clicked.connect(self.adicionar_volta) #Envia Volta
        
        ######################################################################################################################
        #Funções da Pagina Consulta
        self.btn_consulta.clicked.connect(self.consultar_viagens)
        self.btn_consulta.clicked.connect(self.consultar_barcos)#Faz com que carregue os dados quando o botão for clicado
        
        ######################################################################################################################
        #Funções da Pagina cadastro
        self.btn_cadastro_piloto.clicked.connect(self.cadastro_piloto) #Botão cadastro pilotos
        self.btn_cadastro_barco.clicked.connect(self.cadastro_barcos) #Botão cadastro barcos
          
    def leftMenu(self):
        
        width = self.left_container.width()
        
        if width == 9:
            newWidth = 200
        else:
            newWidth = 9
            
        self.animation = QtCore.QPropertyAnimation(self.left_container, b"maximumWidth")
        self.animation.setDuration(500)
        self.animation.setStartValue(width)
        self.animation.setEndValue(newWidth)
        self.animation.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
        self.animation.start()
    
    def load_workbook(self, filename):
        self.workbook = openpyxl.load_workbook(filename)
        self.barcos_sheet = self.workbook["Barcos"]
        self.pilotos_sheet = self.workbook["Pilotos"]
        self.viagens_sheet = self.workbook["Viagens"] 
        self.clientes_sheet = self.workbook["Clientes"] 
     
    def carregar_barcos_combobox(self):
        rows = list(self.barcos_sheet.iter_rows(values_only=True))
        barcos = [row[1] for row in rows[1:] if row[1] is not None]
        self.barcos_model.clear()
        for barco in barcos:
            item = QStandardItem(barco)
            self.barcos_model.appendRow(item)

    def carregar_pilotos_combobox(self):
        rows = list(self.pilotos_sheet.iter_rows(values_only=True))
        pilotos = [row[0] for row in rows[1:] if row[0] is not None]
        self.pilotos_model.clear()  
        for piloto in pilotos:
            item = QStandardItem(piloto)
            self.pilotos_model.appendRow(item)
    
    def adicionar_viagem(self):
        barco_selecionado = self.barcos_comboBox.currentText()
        piloto_selecionado = self.pilotos_comboBox.currentText()
        gasolina_saida = self.qtn_saida_lineEdit.text()
        gasolina_volta = None
        hora_volta = None
        minutos = None
        nome_cliente = self.cliente_nome_lineEdit.text()
        telefone_cliente = self.telefone_cliente_lineEdit.text()
        gasolina_total = 0
        cor_personalizada = QColor(0, 0, 0) 
        
        if not gasolina_saida.isnumeric():
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText(f"Por favor no campo Gasolina digitar somente numeros!")
            msg.setStyleSheet(f"QLabel {{ color: {cor_personalizada.name()}; }}")
            msg.setWindowTitle("Aviso!")
            msg.exec_()
            return

        if not barco_selecionado or not piloto_selecionado:
            QMessageBox.warning(self, "Erro", "Preencha todos os campos.")
            return

        data_hora_atual = datetime.now()
        data = data_hora_atual.strftime("%d/%m/%Y")
        hora = data_hora_atual.strftime("%H:%M:%S")

        viagem_id = len(list(self.viagens_sheet["A"])) + 1

        self.viagens_sheet.append(
            [viagem_id, barco_selecionado, piloto_selecionado, gasolina_saida, gasolina_volta, data, hora, hora_volta,minutos,'Aberta',gasolina_total,nome_cliente if nome_cliente is not None else 'Não Informado', telefone_cliente if telefone_cliente is not None else 'Não Informado']
        )

        
        
        self.workbook.save("controle_de_barcos.xlsx")
        self.qtn_saida_lineEdit.clear()
        self.qtn_saida_lineEdit.setFocus()
        self.telefone_cliente_lineEdit.clear()
        self.cliente_nome_lineEdit.clear()
        self.cpf_cliente_lineEdit.clear()
        
        cor_personalizada = QColor(0, 0, 0)  

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(f"Seu Código é: {viagem_id}")

       
        msg.setStyleSheet(f"QLabel {{ color: {cor_personalizada.name()}; }}")

        msg.setWindowTitle("Código da Viagem")
        msg.exec_()

    def adicionar_cliente(self):
        
        cliente_id = len(list(self.clientes_sheet["A"])) + 1
        nome = self.cliente_nome_lineEdit.text()
        cpf = self.cpf_cliente_lineEdit.text()
        telefone = self.telefone_cliente_lineEdit.text()
        

        self.clientes_sheet.append(
        [cliente_id,nome,cpf if cpf else 'Pendente',telefone if cpf else 'Pendente'])

        
        self.workbook.save('controle_de_barcos.xlsx')
              
    def adicionar_volta(self):
        id_viagem = self.id_lineEdit.text()
        gasolina_volta = self.qtn_volta_lineEdit.text()
        cor_personalizada = QColor(0, 0, 0)

        if not gasolina_volta.isnumeric():
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText(f"Por favor no campo Gasolina digitar somente numeros!")
            msg.setStyleSheet(f"QLabel {{ color: {cor_personalizada.name()}; }}")
            msg.setWindowTitle("Aviso!")
            msg.exec_()
            return
        
        gasolina_volta = float(gasolina_volta)
        
        if not id_viagem :
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText(f"Preencha todos os campos! ID ")

            msg.setStyleSheet(f"QLabel {{ color: {cor_personalizada.name()}; }}")

            msg.setWindowTitle("ERRO!")
            msg.exec_()
            return

        for row in self.viagens_sheet.iter_rows(values_only=True):
            if str(row[0]) == id_viagem:

                if row[9] == 'Finalizada':
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Information)
                    msg.setText(f"Esta viagem já está finalizada e não pode ser alterada.")

                    msg.setStyleSheet(f"QLabel {{ color: {cor_personalizada.name()}; }}")

                    msg.setWindowTitle("Aviso!")
                    msg.exec_()

                    return

                hora_volta = datetime.now().strftime("%H:%M:%S") 
                
                hora_saida = datetime.strptime(row[6], "%H:%M:%S") 
                hora_volta = datetime.strptime(hora_volta, "%H:%M:%S") 
                diferenca = hora_volta - hora_saida
                segundos_viagem = int(diferenca.total_seconds()) 
       
                minutos_viagem = round(segundos_viagem / 60)
                       
                gasolina_saida = float(row[3])
                
                gasolina = 0
                
                if gasolina_saida > gasolina_volta:
                    gasolina = gasolina_saida - gasolina_volta
                  
                self.viagens_sheet.cell(row=row[0], column=5, value=gasolina_volta)
                self.viagens_sheet.cell(row=row[0], column=8, value=hora_volta.strftime("%H:%M:%S"))
                self.viagens_sheet.cell(row=row[0], column=9, value=minutos_viagem)  
                self.viagens_sheet.cell(row=row[0], column=11, value=gasolina)
               
                barco_selecionado = row[1]

                linha_barco = None
                
                for index, barcos_row in enumerate(self.barcos_sheet.iter_rows(values_only=True), start=1):
                    if barcos_row[1] == barco_selecionado:
                    
                        valor_existente = barcos_row[3] if barcos_row[3] else 0  # Se não houver valor, comece com 0
                        novo_valor = valor_existente + minutos_viagem
                        
                        gasolina_existente = barcos_row[5] if barcos_row[5] else 0
                        
                        gasolina_total = gasolina_existente + gasolina
                        linha_barco = index

                        break 

                        
                if linha_barco is not None:         
                    self.barcos_sheet.cell(row=linha_barco, column=4, value=novo_valor)
                    self.barcos_sheet.cell(row=linha_barco, column=6, value=gasolina_total)
                    self.workbook.save("controle_de_barcos.xlsx")
                    
                    self.load_workbook('./BKP/Planilha_Base/controle_de_barcos.xlsx')
                    self.barcos_sheet = self.workbook['Barcos']
                    self.barcos_sheet.cell(row=linha_barco, column=4, value=novo_valor)
                    self.barcos_sheet.cell(row=linha_barco, column=6, value=gasolina_total)
                    self.workbook.save('./BKP/Planilha_Base/controle_de_barcos.xlsx')
                    
                    self.load_workbook('controle_de_barcos.xlsx')
                    self.barcos_sheet = self.workbook['Barcos']
               
                #if gasolina_volta is not None:
                self.viagens_sheet.cell(row=row[0], column=10, value=str('Finalizada'))

                self.workbook.save("controle_de_barcos.xlsx")
                
                cor_personalizada = QColor(0, 0, 0)
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setText(f"Viagem Finalizada com sucesso!")
                msg.setStyleSheet(f"QLabel {{ color: {cor_personalizada.name()}; }}")
                msg.setWindowTitle("Código da Viagem")
                msg.exec_()
                
                self.qtn_volta_lineEdit.clear()
                self.id_lineEdit.clear()
                return

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(f"ID Inválido")
        msg.setStyleSheet(f"QLabel {{ color: {cor_personalizada.name()}; }}")
        msg.setWindowTitle("Aviso!")
        msg.exec_()
        
    def consultar_viagens(self):

        viagens_text = self.text_consulta_viagens
        viagens_text.clear() 
        viagens_text.setReadOnly(True)

        viagens = list(self.viagens_sheet.iter_rows(values_only=True))[1:]  # Ignora a primeira linha (cabeçalho)
        viagens = [viagem for viagem in viagens if viagem[0] is not None]

        # Inverte como é exibido no consulta
        #viagens.sort(key=lambda viagem: int(viagem[0]), reverse=True)

        for viagem in viagens:
            viagens_text.append(
                f"ID: {viagem[0]}, Barco: {viagem[1]}, Piloto: {viagem[2]}, Gasolina: {viagem[3]}, Gasolina Volta: {viagem[4] if viagem[4] else 'Pendente'}, Data: {viagem[5]}, Hora: {viagem[6]}, Data Volta: {viagem[7] if viagem[7] else 'Pendente'}, Nome do Cliente: {viagem[11] if viagem[11] else 'Pendente'},  Telefone do Cliente: {viagem[12] if viagem[12] else 'Pendente'}\n\n"
            ) 

    def consultar_barcos(self):
        consulta_text = self.consulta_barcos
        consulta_text.clear()
        consulta_text.setReadOnly(True)

        consultas = list(self.barcos_sheet.iter_rows(values_only=True))[1:]  # Ignora a primeira linha (cabeçalho)
        consulta = [consulta for consulta in consultas if consulta[0] is not None]

        for consulta_linha in consulta:
            horas_uso = consulta_linha[3]  
            horas = horas_uso // 60
            minutos = horas_uso % 60

            consulta_text.append(
                f"Barco: {consulta_linha[1]}\nHoras de Uso: {horas:02d}:{minutos:02d}\n\n"
            )
            
    def cadastro_piloto(self):
       
        self.aba_pilotos = self.workbook['Pilotos']
        
            
        self.proxima_linha = self.aba_pilotos.max_row + 1

        
        nome = self.nomelineEdit.text()
        cpf = self.cpf_lineEdit.text()

        self.aba_pilotos.cell(row=self.proxima_linha, column=1, value=nome)
        self.aba_pilotos.cell(row=self.proxima_linha, column=2, value=cpf)
        self.workbook.save('controle_de_barcos.xlsx')
        
        
        self.load_workbook('./BKP/Planilha_Base/controle_de_barcos.xlsx')
        
        self.aba_pilotos = self.workbook['Pilotos']
        self.aba_pilotos.cell(row=self.proxima_linha, column=1, value=nome)
        self.aba_pilotos.cell(row=self.proxima_linha, column=2, value=cpf)
        self.workbook.save('./BKP/Planilha_Base/controle_de_barcos.xlsx')
       
        item = QStandardItem(nome)
        self.pilotos_model.appendRow(item)

       
        self.nomelineEdit.clear()
        self.cpf_lineEdit.clear()

        
        self.workbook.close()
    
    def cadastro_barcos(self):
 
        self.aba_barcos = self.workbook['Barcos']

      
        self.proxima_linha = self.aba_barcos.max_row + 1
        
        barco_id = len(list(self.barcos_sheet["A"])) + 1

        
        modelo = self.modelo_lineEdit.text()
        motor = self.motor_lineEdit.text()
        contador = 0
        gasolina = 0

        self.aba_barcos.cell(row=self.proxima_linha, column=1, value=barco_id)
        self.aba_barcos.cell(row=self.proxima_linha, column=2, value=modelo)
        self.aba_barcos.cell(row=self.proxima_linha, column=3, value=motor)
        self.aba_barcos.cell(row=self.proxima_linha, column=4, value=0)
        self.aba_barcos.cell(row=self.proxima_linha, column=5, value=contador)
        self.aba_barcos.cell(row=self.proxima_linha, column=6, value=gasolina)
        self.workbook.save('controle_de_barcos.xlsx')
        
        
        self.load_workbook('./BKP/Planilha_Base/controle_de_barcos.xlsx')
        self.aba_barcos = self.workbook['Barcos']
        self.aba_barcos.cell(row=self.proxima_linha, column=1, value=barco_id)
        self.aba_barcos.cell(row=self.proxima_linha, column=2, value=modelo)
        self.aba_barcos.cell(row=self.proxima_linha, column=3, value=motor)
        self.aba_barcos.cell(row=self.proxima_linha, column=4, value=0)
        self.aba_barcos.cell(row=self.proxima_linha, column=5, value=contador)
        self.aba_barcos.cell(row=self.proxima_linha, column=6, value=gasolina)
        self.workbook.save('./BKP/Planilha_Base/controle_de_barcos.xlsx')
        
        item = QStandardItem(modelo)
        self.barcos_model.appendRow(item)

        self.modelo_lineEdit.clear()
        self.motor_lineEdit.clear()
        self.workbook.close()
    
if __name__ =="__main__":
    app = QApplication(sys.argv)
    #window = MainWindow() #Inicia a tela principal
    window_login = Login()
    window_login.show()
    #window.show()  #Inicia a tela principal
    app.exec()
    